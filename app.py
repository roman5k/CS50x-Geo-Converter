import pandas as pd
from cs50 import SQL
from flask import Flask, flash, redirect, render_template, request, session, send_file, jsonify
from flask_session import Session
from werkzeug.security import check_password_hash, generate_password_hash
from helpers import apology, login_required
from pyproj import Transformer
from io import BytesIO
from openpyxl import load_workbook

# Configure application
app = Flask(__name__)

# Configure session to use filesystem (instead of signed cookies)
app.config["SESSION_PERMANENT"] = False
app.config["SESSION_TYPE"] = "filesystem"
app.config['UPLOAD_FOLDER'] = 'uploads'
Session(app)

# Configure CS50 Library to use SQLite database
db = SQL("sqlite:///geo.db")

@app.after_request
def after_request(response):
    """Ensure responses aren't cached"""
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Expires"] = 0
    response.headers["Pragma"] = "no-cache"
    return response

@app.route("/")
def index():
    return render_template("index.html")

@app.route('/convert', methods=["GET", "POST"])
@login_required
def convert():
    # cs50 duck and gpt-4o helped me with this.
    if request.method == "POST":
        if 'file' not in request.files:
            return apology("No file part", 400)
        file = request.files['file']
        if file.filename == '':
            return apology("No selected file", 400)

        data = request.form
        longitudeColumn = data['longitudeColumn']
        latitudeColumn = data['latitudeColumn']
        coordSys = data['coordSys']
        action = data.get('action', 'preview')

        # Load the workbook and get the first sheet
        wb = load_workbook(file)
        ws = wb.active

        # Extract the relevant columns
        headers = [cell.value for cell in ws[1]]
        longitude_index = headers.index(longitudeColumn)
        latitude_index = headers.index(latitudeColumn)
        coordinates = [(row[longitude_index].value, row[latitude_index].value) for row in ws.iter_rows(min_row=2)]

        # Transform the coordinates
        # help from pyproj documentation
        transformer = Transformer.from_crs(coordSys, "EPSG:4326", always_xy=True)
        transformed_coordinates = [transformer.transform(lon, lat) for lon, lat in coordinates]
        formatted_coordinates = [(f"{lon:.5f}", f"{lat:.5f}") for lon, lat in transformed_coordinates]

        if action == 'preview':
            # Return only the first two transformed coordinates as a plain string
            first_two_coordinates = formatted_coordinates[:2]
            return f"{first_two_coordinates[0][0]}, {first_two_coordinates[0][1]}"
        elif action == 'convert':
            # Update the original data with transformed coordinates
            for i, (lon, lat) in enumerate(formatted_coordinates):
                ws.cell(row=i + 2, column=longitude_index + 1, value=lon)
                ws.cell(row=i + 2, column=latitude_index + 1, value=lat)

            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return send_file(output, as_attachment=True, download_name='converted_coordinates.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    else:
        return render_template("convert.html")

@app.route("/settings", methods=["GET", "POST"])
@login_required
def settings():
    # Redirect user to login form
    return render_template("/settings.html")

# cs50 duck and gpt-4o helped me with this.
@app.route("/change_password", methods=["POST"])
@login_required
def change_password():
    user_id = session.get("user_id")
    if not user_id:
        return redirect("/login")

    old_password = request.form.get("old_password")
    new_password = request.form.get("new_password")
    confirm_password = request.form.get("confirm_password")

    if not old_password or not new_password or not confirm_password:
        return apology("All fields are required", 400)

    if new_password != confirm_password:
        return apology("New passwords do not match", 400)

    # Query database for the user's current password hash
    rows = db.execute("SELECT hash FROM users WHERE id = ?", user_id)
    if len(rows) != 1 or not check_password_hash(rows[0]["hash"], old_password):
        return apology("Invalid current password", 400)

    # Update the password hash in the database
    new_hash = generate_password_hash(new_password)
    db.execute("UPDATE users SET hash = ? WHERE id = ?", new_hash, user_id)

    flash("Password changed successfully")
    return redirect("/settings")

# cs50 duck and gpt-4o helped me with this.
@app.route("/delete_account", methods=["POST"])
@login_required
def delete_account():
    user_id = session.get("user_id")
    if not user_id:
        return redirect("/login")

    password = request.form.get("password")

    if not password:
        return apology("Password is required", 400)

    # Query database for the user's current password hash
    rows = db.execute("SELECT hash FROM users WHERE id = ?", user_id)
    if len(rows) != 1 or not check_password_hash(rows[0]["hash"], password):
        return apology("Invalid password", 400)

    # Delete the user from the database
    db.execute("DELETE FROM users WHERE id = ?", user_id)

    # Clear the session
    session.clear()

    flash("Account deleted successfully")
    return redirect("/")


@app.route("/login", methods=["GET", "POST"])
def login():
    """Log user in"""
    # Forget any user_id
    session.clear()

    # User reached route via POST (as by submitting a form via POST)
    if request.method == "POST":
        # Ensure username was submitted
        if not request.form.get("username"):
            return apology("must provide username", 403)

        # Ensure password was submitted
        elif not request.form.get("password"):
            return apology("must provide password", 403)

        # Query database for username
        rows = db.execute(
            "SELECT * FROM users WHERE username = ?", request.form.get("username")
        )

        # Ensure username exists and password is correct
        if len(rows) != 1 or not check_password_hash(
            rows[0]["hash"], request.form.get("password")
        ):
            return apology("invalid username and/or password", 403)

        # Remember which user has logged in
        session["user_id"] = rows[0]["id"]

        # Redirect user to home page
        return redirect("/convert")

    # User reached route via GET (as by clicking a link or via redirect)
    else:
        return render_template("login.html")

@app.route("/logout")
@login_required
def logout():
    """Log user out"""
    # Forget any user_id
    session.clear()

    # Redirect user to login form
    return redirect("/")

@app.route("/register", methods=["GET", "POST"])
def register():
    """Register user"""
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        confirmation = request.form.get("confirmation")

        if not username or not password or not confirmation:
            return apology("THE FIELDS CANNOT BE EMPTY")
        elif password != confirmation:
            return apology("YOUR ENTERED PASSWORDS DO NOT MATCH.")
        elif username and password == confirmation:
            try:
                hash = generate_password_hash(password)
                db.execute(
                    "INSERT INTO users (username, hash) "
                    "VALUES (?, ?); ",
                    username,
                    hash
                )
                # Query database for username
                rows = db.execute(
                    "SELECT * FROM users WHERE username = ?", username
                )
                # Remember which user has logged in
                session["user_id"] = rows[0]["id"]
                # Redirect user to home page
                return redirect("/convert")
            except:
                return apology("THIS USERNAME IS ALREADY IN USE.")
    return render_template("register.html")
